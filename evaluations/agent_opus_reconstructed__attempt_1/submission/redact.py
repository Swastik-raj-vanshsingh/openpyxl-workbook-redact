# Copyright (c) 2010-2024 openpyxl

"""
Produce a redacted copy of a workbook for sharing with an outside party.

The :func:`redact_workbook` function copies a single workbook, dropping the
sheets and columns named in a spec.  :func:`redact_directory` runs the same
operation over a folder and writes a CSV report; it is also available as::

    python -m openpyxl.redact INPUT_DIR OUTPUT_DIR --spec SPEC.json --report REPORT.csv

Removing a column shifts everything to its right one column to the left.
openpyxl deliberately does not manage dependencies -- formulae, defined names,
tables, filters, merged ranges, validations, conditional formats, charts,
column widths -- when columns are deleted (see
``doc/editing_worksheets.rst``), so everything that refers to a cell by
position is rewritten here.

Two rules drive the rewriting:

* a reference that does *not* touch removed data is re-pointed at the cells it
  always meant, so the formula keeps computing the same answer;
* a reference that *does* touch removed data cannot be repaired -- whatever
  slid into its place is a different number -- so the formula is destroyed and
  the cell shows the spec's marker instead.

The same two rules run over defined names, tables, filters, merged ranges,
validations, conditional formats, print areas, panes and charts.  Anything the
copy could carry a removed value in but cannot be repaired -- a pivot cache, a
scenario, a filter on a removed column -- is dropped rather than kept.
References out of the workbook go the same way: external links are stripped and
the formulas and names that used them are marked, so the copy stands alone.

The report row for a workbook records:

``status``
    ``redacted`` if a sheet or a column went, ``unchanged`` if nothing in the
    spec matched, ``failed`` if no copy could be written.
``sheets_removed``
    how many sheets the spec named and the workbook had.
``columns_removed``
    how many columns went from the sheets that stayed.
``cells_removed``
    how many cells holding a value went with them, counting the cells of a
    removed sheet.
``detail``
    what went, in the spec's own words, plus counts of the dependent things
    that had to go with it.  It never carries a removed value.

Known limits: a formula that builds a reference at run time -- ``INDIRECT``,
``OFFSET`` with a computed column -- cannot be followed, so such a formula is
left alone and may read a column that moved.  A structured reference is
followed by name rather than by position, which is how Excel resolves it.
"""

import argparse
import csv
import datetime
import json
import os
import re
import sys
import tempfile
from bisect import bisect_left

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.formula.tokenizer import Tokenizer, Token, TokenizerError
from openpyxl.packaging.core import DocumentProperties
from openpyxl.packaging.custom import CustomPropertyList
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.print_settings import ColRange, PrintArea
from openpyxl.worksheet.scenario import ScenarioList
from openpyxl.formatting.formatting import (
    ConditionalFormatting,
    ConditionalFormattingList,
)


__all__ = [
    "SpecError",
    "RedactionError",
    "REPORT_FIELDS",
    "STATUS_REDACTED",
    "STATUS_UNCHANGED",
    "STATUS_FAILED",
    "load_spec",
    "validate_spec",
    "redact_workbook",
    "redact_directory",
    "main",
]


REPORT_FIELDS = ("workbook", "status", "sheets_removed", "columns_removed",
                 "cells_removed", "detail")

STATUS_REDACTED = "redacted"
STATUS_UNCHANGED = "unchanged"
STATUS_FAILED = "failed"

WORKBOOK_SUFFIX = ".xlsx"

#: The four fields a spec may contain, and the type each one must have.
SPEC_FIELDS = {
    "drop_sheets": list,
    "drop_columns": list,
    "marker": str,
    "organisation": str,
}


class SpecError(ValueError):
    """The redaction spec is not shaped the way the tool requires."""


class RedactionError(Exception):
    """A workbook could not be redacted; no copy has been written for it."""


def _key(value):
    """Fold a sheet, header or name to its comparison form."""
    if value is None:
        return None
    if not isinstance(value, str):
        value = str(value)
    return value.strip().casefold()


###############################################################################
# The spec
###############################################################################

def validate_spec(spec):
    """
    Check that `spec` is shaped the way the ticket describes and return a
    normalised copy of it.

    :raises SpecError: if a field is missing, unknown, or the wrong type.
    """
    if not isinstance(spec, dict):
        raise SpecError("spec must be a JSON object, not "
                        f"{type(spec).__name__}")

    unknown = sorted(set(spec) - set(SPEC_FIELDS))
    if unknown:
        raise SpecError(f"spec contains unknown field(s): {', '.join(unknown)}")

    missing = sorted(set(SPEC_FIELDS) - set(spec))
    if missing:
        raise SpecError(f"spec is missing field(s): {', '.join(missing)}")

    for field, expected in SPEC_FIELDS.items():
        value = spec[field]
        # bool is an int, not a str or list, but be explicit about it anyway
        if isinstance(value, bool) or not isinstance(value, expected):
            raise SpecError(f"spec field {field!r} must be a "
                            f"{'list' if expected is list else 'string'}")

    for field in ("drop_sheets", "drop_columns"):
        for item in spec[field]:
            if not isinstance(item, str):
                raise SpecError(f"every entry in {field!r} must be a string")

    if not spec["marker"]:
        raise SpecError("spec field 'marker' must not be empty")
    if not spec["organisation"]:
        raise SpecError("spec field 'organisation' must not be empty")

    return {
        "drop_sheets": list(spec["drop_sheets"]),
        "drop_columns": list(spec["drop_columns"]),
        "marker": spec["marker"],
        "organisation": spec["organisation"],
    }


def load_spec(path):
    """Read and validate a spec from a JSON file."""
    try:
        with open(path, encoding="utf-8") as src:
            raw = json.load(src)
    except OSError as e:
        raise SpecError(f"cannot read spec {path}: {e}") from e
    except ValueError as e:
        raise SpecError(f"spec {path} is not valid JSON: {e}") from e
    return validate_spec(raw)


###############################################################################
# Column maps
###############################################################################

class _SheetPlan:
    """Which columns go from one sheet, and where the survivors end up."""

    def __init__(self, title, removed=()):
        self.title = title
        self.key = _key(title)
        self.removed = sorted(removed)
        self._removed = set(self.removed)

    @property
    def changed(self):
        return bool(self.removed)

    def is_removed(self, col):
        return col in self._removed

    def map_col(self, col):
        """New index of `col`, or None if the column has gone."""
        if col in self._removed:
            return None
        return col - bisect_left(self.removed, col)

    def any_removed_between(self, lo, hi):
        i = bisect_left(self.removed, lo)
        return i < len(self.removed) and self.removed[i] <= hi

    def first_kept(self, lo, hi):
        col = lo
        for r in self.removed[bisect_left(self.removed, lo):]:
            if r > col:
                break
            col = r + 1
        return col if col <= hi else None

    def last_kept(self, lo, hi):
        col = hi
        for r in reversed(self.removed[:bisect_left(self.removed, hi + 1)]):
            if r < col:
                break
            col = r - 1
        return col if col >= lo else None


class _TablePlan:
    """What happened to one worksheet table."""

    def __init__(self, name, removed=False, dropped_columns=()):
        self.name = name
        self.removed = removed
        self.dropped_columns = set(dropped_columns)

    @property
    def changed(self):
        return self.removed or bool(self.dropped_columns)


class _Poisoned(Exception):
    """A reference cannot survive the redaction."""


###############################################################################
# Reference rewriting
###############################################################################

_ROW_RANGE_RE = re.compile(r"^(\$?)([1-9][0-9]{0,6}):(\$?)([1-9][0-9]{0,6})$")
_COL_RANGE_RE = re.compile(r"^(\$?)([A-Za-z]{1,3}):(\$?)([A-Za-z]{1,3})$")
_CELL_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)([1-9][0-9]{0,6})$")
_STRUCTURED_RE = re.compile(r"^(?P<name>[^\[\]!]+)\[(?P<body>.*)\]$", re.S)
_BRACKETED_RE = re.compile(r"\[([^\[\]]*)\]")
# `A1:INDEX(` and friends arrive from the tokenizer as a single FUNC token
_FUNC_WITH_REF_RE = re.compile(r"^(?P<ref>.+):(?P<func>[^:()]+)\($")


def _unquote_sheet(name):
    if len(name) > 1 and name.startswith("'") and name.endswith("'"):
        return name[1:-1].replace("''", "'")
    return name


class _Rewriter:
    """
    Rewrites the references in a formula so they still mean what they meant.

    Every reference is remapped through the column map of the sheet it points
    at, so each one moves by its own amount -- unlike
    :class:`openpyxl.formula.translate.Translator`, which shifts a whole
    formula by a single offset.
    """

    def __init__(self, plan):
        self.plan = plan

    # -- entry points ------------------------------------------------------

    def formula(self, text, sheet_key):
        """Rewrite a complete formula (leading '=' included)."""
        if not isinstance(text, str) or not text.startswith("="):
            return text
        try:
            tokens = Tokenizer(text).items
        except (TokenizerError, IndexError, ValueError) as e:
            raise _Poisoned(f"unparsable formula: {e}") from e
        if not tokens:
            return text
        if tokens[0].type == Token.LITERAL:
            return text

        out = ["="]
        for token in tokens:
            if token.type == Token.OPERAND and token.subtype == Token.RANGE:
                out.append(self.reference(token.value, sheet_key))
            elif token.type == Token.FUNC and token.subtype == Token.OPEN:
                m = _FUNC_WITH_REF_RE.match(token.value)
                if m:
                    out.append(self.reference(m.group("ref"), sheet_key))
                    out.append(":" + m.group("func") + "(")
                else:
                    out.append(token.value)
            else:
                out.append(token.value)
        return "".join(out)

    def body(self, text, sheet_key):
        """Rewrite a formula stored without its leading '=' (rules, tables)."""
        if not isinstance(text, str) or not text:
            return text
        return self.formula("=" + text, sheet_key)[1:]

    # -- one reference -----------------------------------------------------

    def reference(self, value, sheet_key):
        sheet_part, rest = "", value
        if "!" in value:
            sheet_part, rest = value.rsplit("!", 1)

        if sheet_part:
            if "[" in sheet_part:
                raise _Poisoned("reference to another workbook")
            title = _unquote_sheet(sheet_part)
            if ":" in title:
                # a 3D reference; the sheets it spans cannot be resolved
                raise _Poisoned("reference spanning several sheets")
            target = self.plan.sheets.get(_key(title))
            if target is None:
                if _key(title) in self.plan.dropped_sheets:
                    raise _Poisoned("reference to a removed sheet")
                raise _Poisoned(f"reference to unknown sheet {title!r}")
            prefix = sheet_part + "!"
        else:
            if rest.startswith("["):
                raise _Poisoned("reference to another workbook")
            # a workbook-scoped name has no sheet of its own; such a value can
            # only name another name, never a bare cell
            target = self.plan.sheets.get(sheet_key)
            prefix = ""

        if "[" in rest:
            self._check_structured(rest)
            return value

        return prefix + self._remap_cells(rest, target, sheet_key)

    def _check_structured(self, rest):
        m = _STRUCTURED_RE.match(rest)
        if m is None:
            raise _Poisoned("unrecognised structured reference")
        table = self.plan.tables.get(_key(m.group("name")))
        if table is None or table.removed:
            raise _Poisoned("reference to a removed table")
        if not table.changed:
            return
        body = m.group("body")
        items = _BRACKETED_RE.findall(body) or ([body] if body else [])
        named = [i for i in items if not i.startswith("#")]
        if not named:
            # the whole table, which is no longer the same table
            raise _Poisoned("reference to a table that lost a column")
        for item in named:
            if _key(item) in table.dropped_columns:
                raise _Poisoned("reference to a removed table column")

    def _remap_cells(self, rest, target, sheet_key):
        m = _ROW_RANGE_RE.match(rest)
        if m is not None:
            self._require(target)
            # whole rows: they covered the removed columns too
            if target.changed:
                raise _Poisoned("reference to whole rows of a changed sheet")
            return rest

        m = _COL_RANGE_RE.match(rest)
        if m is not None:
            self._require(target)
            lo = column_index_from_string(m.group(2).upper())
            hi = column_index_from_string(m.group(4).upper())
            a, b = min(lo, hi), max(lo, hi)
            if target.any_removed_between(a, b):
                raise _Poisoned("reference covering a removed column")
            return (f"{m.group(1)}{get_column_letter(target.map_col(lo))}:"
                    f"{m.group(3)}{get_column_letter(target.map_col(hi))}")

        if ":" in rest:
            pieces = rest.split(":")
            matches = [_CELL_RE.match(p) for p in pieces]
            if all(matches):
                self._require(target)
                cols = [column_index_from_string(mm.group(2).upper())
                        for mm in matches]
                if target.any_removed_between(min(cols), max(cols)):
                    raise _Poisoned("reference covering a removed column")
                return ":".join(self._remap_cell(mm, target)
                                for mm in matches)
            # at least one endpoint is a defined name
            for piece, mm in zip(pieces, matches):
                if mm is None:
                    self._check_name(piece, sheet_key)
            if target is not None and target.changed:
                raise _Poisoned("range anchored on a name in a changed sheet")
            return rest

        m = _CELL_RE.match(rest)
        if m is not None:
            self._require(target)
            col = column_index_from_string(m.group(2).upper())
            if target.is_removed(col):
                raise _Poisoned("reference to a removed column")
            return self._remap_cell(m, target)

        self._check_name(rest, sheet_key)
        return rest

    @staticmethod
    def _require(target):
        if target is None:
            raise _Poisoned("cell reference with no sheet to resolve it")

    @staticmethod
    def _remap_cell(m, target):
        col = column_index_from_string(m.group(2).upper())
        return (f"{m.group(1)}{get_column_letter(target.map_col(col))}"
                f"{m.group(3)}{m.group(4)}")

    def _check_name(self, name, sheet_key):
        key = _key(name)
        table = self.plan.tables.get(key)
        if table is not None and table.changed:
            raise _Poisoned("reference to a table that changed")
        if (sheet_key, key) in self.plan.poisoned_names:
            raise _Poisoned(f"reference to removed name {name!r}")
        if (None, key) in self.plan.poisoned_names:
            raise _Poisoned(f"reference to removed name {name!r}")


###############################################################################
# The plan for a whole workbook
###############################################################################

class _Plan:

    def __init__(self):
        self.dropped_sheets = set()      # casefolded titles
        self.sheets = {}                 # casefolded title -> _SheetPlan
        self.tables = {}                 # casefolded name -> _TablePlan
        self.poisoned_names = set()      # (sheet key or None, name key)

    @property
    def any_columns_removed(self):
        return any(sp.changed for sp in self.sheets.values())

    @property
    def changed(self):
        return bool(self.dropped_sheets) or self.any_columns_removed


###############################################################################
# Redacting one workbook
###############################################################################

class _WorkbookRedactor:

    def __init__(self, wb, spec):
        self.wb = wb
        self.spec = spec
        self.marker = spec["marker"]
        self.organisation = spec["organisation"]
        # the report names what went using the spec's own spelling, trimmed
        self.drop_sheets = {_key(s): s.strip()
                            for s in spec["drop_sheets"] if _key(s)}
        self.drop_columns = {_key(c): c.strip()
                             for c in spec["drop_columns"] if _key(c)}
        self.plan = _Plan()
        self.rewriter = _Rewriter(self.plan)

        self.sheets_removed = 0
        self.columns_removed = 0
        self.cells_removed = 0
        self.removed_sheet_names = []
        self.removed_column_names = {}   # spec spelling -> [sheet titles]
        self.dropped_headers = set()     # header keys actually removed
        self.counts = {}

    def _bump(self, what, amount=1):
        if amount:
            self.counts[what] = self.counts.get(what, 0) + amount

    # -- planning ----------------------------------------------------------

    def plan_workbook(self):
        doomed = [s for s in self.wb._sheets if _key(s.title) in self.drop_sheets]
        survivors = [s for s in self.wb._sheets if s not in doomed]

        if doomed and not survivors:
            raise RedactionError(
                "the spec would remove every sheet in the workbook")
        if doomed and not any(s.sheet_state == "visible" for s in survivors):
            raise RedactionError(
                "the spec would leave the workbook with no visible sheet")

        for sheet in doomed:
            self.plan.dropped_sheets.add(_key(sheet.title))
            self.removed_sheet_names.append(self.drop_sheets[_key(sheet.title)])
        self.sheets_removed = len(doomed)

        for sheet in survivors:
            if not isinstance(sheet, Worksheet):
                self.plan.sheets[_key(sheet.title)] = _SheetPlan(sheet.title)
                continue
            removed = []
            if self.drop_columns and sheet.max_row >= 1:
                for col in range(1, sheet.max_column + 1):
                    cell = sheet._cells.get((1, col))
                    if cell is None:
                        continue
                    name = self.drop_columns.get(_key(cell.value))
                    if name is None:
                        continue
                    removed.append(col)
                    self.dropped_headers.add(_key(cell.value))
                    self.removed_column_names.setdefault(name, [])
                    if sheet.title not in self.removed_column_names[name]:
                        self.removed_column_names[name].append(sheet.title)
            self.plan.sheets[_key(sheet.title)] = _SheetPlan(sheet.title, removed)

        self.columns_removed = sum(len(sp.removed)
                                   for sp in self.plan.sheets.values())

        self._plan_tables(doomed, survivors)
        self._count_removed_cells(doomed, survivors)
        self._plan_names()

    def _plan_tables(self, doomed, survivors):
        for sheet in doomed:
            # TableList.items() yields refs, not tables; go through the values
            for table in getattr(sheet, "_tables", {}).values():
                self.plan.tables[_key(table.name)] = _TablePlan(table.name,
                                                                removed=True)

        for sheet in survivors:
            if not isinstance(sheet, Worksheet):
                continue
            sp = self.plan.sheets[_key(sheet.title)]
            for table in sheet._tables.values():
                cr = CellRange(table.ref)
                dropped = set()
                for offset, column in enumerate(table.tableColumns):
                    col = cr.min_col + offset
                    if sp.is_removed(col):
                        dropped.add(_key(column.name))
                gone = sp.first_kept(cr.min_col, cr.max_col) is None
                self.plan.tables[_key(table.name)] = _TablePlan(
                    table.name, removed=gone, dropped_columns=dropped)

    def _count_removed_cells(self, doomed, survivors):
        for sheet in doomed:
            if not isinstance(sheet, Worksheet):
                continue
            self.cells_removed += sum(1 for c in sheet._cells.values()
                                      if c.value is not None)
        for sheet in survivors:
            if not isinstance(sheet, Worksheet):
                continue
            sp = self.plan.sheets[_key(sheet.title)]
            if not sp.changed:
                continue
            self.cells_removed += sum(1 for (_r, c), cell in sheet._cells.items()
                                      if sp.is_removed(c)
                                      and cell.value is not None)

    def _iter_defined_names(self):
        """Yield (scope key, DefinedNameDict, DefinedName) for every name."""
        for defn in list(self.wb.defined_names.values()):
            yield None, self.wb.defined_names, defn
        for sheet in self.wb._sheets:
            if _key(sheet.title) in self.plan.dropped_sheets:
                continue
            for defn in list(sheet.defined_names.values()):
                yield _key(sheet.title), sheet.defined_names, defn

    def _plan_names(self):
        """
        Work out which defined names cannot survive.  A name may refer to
        another name, so this runs to a fixed point.
        """
        if not self.plan.changed and not self.wb._external_links:
            return
        while True:
            grew = False
            for scope, _owner, defn in self._iter_defined_names():
                token = (scope, _key(defn.name))
                if token in self.plan.poisoned_names:
                    continue
                try:
                    if not defn.value or defn.is_external:
                        raise _Poisoned("name refers outside the workbook")
                    self.rewriter.body(defn.value, scope)
                except (_Poisoned, TypeError, ValueError):
                    self.plan.poisoned_names.add(token)
                    grew = True
            if not grew:
                return

    # -- rewriting ---------------------------------------------------------

    def apply(self):
        # references to another workbook have to go the same way references to
        # a removed column do, so a workbook with external links is rewritten
        # even when nothing in the spec matched it
        rewrite = self.plan.changed or bool(self.wb._external_links)
        self._strip_external_links()
        # pivots first: their rendered output is read in original coordinates
        self._apply_pivots()
        if rewrite:
            self._apply_names()
            for sheet in self.wb._sheets:
                if _key(sheet.title) in self.plan.dropped_sheets:
                    continue
                if isinstance(sheet, Worksheet):
                    self._apply_worksheet(sheet)
                self._apply_charts(sheet)
        self._remove_sheets()
        self._apply_properties()

    def _strip_external_links(self):
        if self.wb._external_links:
            self._bump("external links removed", len(self.wb._external_links))
            self.wb._external_links = []

    def _apply_names(self):
        for scope, owner, defn in self._iter_defined_names():
            if (scope, _key(defn.name)) in self.plan.poisoned_names:
                del owner[defn.name]
                self._bump("defined names removed")
                continue
            try:
                defn.value = self.rewriter.body(defn.value, scope)
            except _Poisoned:                             # pragma: no cover
                del owner[defn.name]
                self._bump("defined names removed")

    def _apply_worksheet(self, ws):
        sp = self.plan.sheets[_key(ws.title)]
        # unmerging first drops the placeholder cells, so deleting a column
        # cannot shuffle them out from under their range
        merges = [str(r) for r in ws.merged_cells.ranges] if sp.changed else []
        for coord in merges:
            ws.unmerge_cells(coord)

        self._apply_cells(ws, sp)
        self._apply_tables(ws, sp)
        self._apply_auto_filter(ws, sp)
        self._apply_validations(ws, sp)
        self._apply_conditional_formats(ws, sp)
        self._apply_print_settings(ws, sp)
        self._apply_scenarios(ws, sp)
        self._apply_views(ws, sp)
        self._apply_breaks(ws, sp)

        if sp.changed:
            self._apply_dimensions(ws, sp)
            for start, length in _runs(sp.removed):
                ws.delete_cols(start, length)

        for coord in merges:
            new = self._shrink_range(coord, sp)
            if new is not None:
                ws.merge_cells(new)
            else:
                self._bump("merged ranges removed")

    # -- cells -------------------------------------------------------------

    def _mark(self, cell):
        """Destroy a cell's formula and show the marker instead."""
        cell._value = self.marker
        cell.data_type = "s"
        self._bump("formulas marked")

    def _apply_cells(self, ws, sp):
        sheet_key = _key(ws.title)
        for (row, col), cell in list(ws._cells.items()):
            if sp.is_removed(col):
                continue
            if cell.hyperlink is not None:
                self._apply_hyperlink(cell, sp, sheet_key)
            if cell.data_type != "f":
                continue
            value = cell._value
            try:
                if isinstance(value, ArrayFormula):
                    ref = self._remap_exact(value.ref, sp)
                    text = self.rewriter.formula(value.text, sheet_key)
                    cell._value = ArrayFormula(ref, text)
                elif isinstance(value, DataTableFormula):
                    value.ref = self._remap_exact(value.ref, sp)
                    for attr in ("r1", "r2"):
                        ref = getattr(value, attr, None)
                        if ref:
                            setattr(value, attr, self._remap_exact(ref, sp))
                else:
                    cell._value = self.rewriter.formula(value, sheet_key)
            except _Poisoned:
                self._mark(cell)

    def _apply_hyperlink(self, cell, sp, sheet_key):
        link = cell.hyperlink
        if link.target:
            cell.hyperlink = None
            self._bump("hyperlinks removed")
            return
        if link.location:
            try:
                link.location = self.rewriter.reference(link.location, sheet_key)
            except _Poisoned:
                cell.hyperlink = None
                self._bump("hyperlinks removed")
                return
        try:
            link.ref = self._remap_exact(link.ref, sp)
        except _Poisoned:                                 # pragma: no cover
            cell.hyperlink = None
            self._bump("hyperlinks removed")

    # -- ranges ------------------------------------------------------------

    def _remap_exact(self, coord, sp):
        """Remap a range that must not lose any of its columns."""
        try:
            cr = CellRange(coord)
        except (ValueError, TypeError) as e:
            raise _Poisoned(f"unusable range {coord!r}") from e
        if sp.any_removed_between(cr.min_col, cr.max_col):
            raise _Poisoned("range covering a removed column")
        cr.min_col = sp.map_col(cr.min_col)
        cr.max_col = sp.map_col(cr.max_col)
        return cr.coord

    def _shrink_range(self, coord, sp):
        """
        Remap a decorative range (a merge, a validation, a format), letting it
        shrink over the columns that went.  None if nothing is left of it.
        """
        if isinstance(coord, CellRange):
            cr = coord
        else:
            try:
                cr = CellRange(str(coord))
            except (ValueError, TypeError):
                # a shape openpyxl cannot model, such as a whole-column filter
                return None
        lo = sp.first_kept(cr.min_col, cr.max_col)
        if lo is None:
            return None
        hi = sp.last_kept(cr.min_col, cr.max_col)
        new = CellRange(min_col=sp.map_col(lo), min_row=cr.min_row,
                        max_col=sp.map_col(hi), max_row=cr.max_row)
        return new.coord

    def _shrink_multi(self, ranges, sp):
        out = []
        for cr in sorted(ranges, key=lambda r: (r.min_col, r.min_row)):
            new = self._shrink_range(cr, sp)
            if new is not None:
                out.append(CellRange(new))
        return out

    # -- worksheet furniture ----------------------------------------------

    def _apply_dimensions(self, ws, sp):
        holder = ws.column_dimensions
        old = list(holder.values())
        holder.clear()
        for dim in old:
            dim.reindex()
            lo = sp.first_kept(dim.min, dim.max)
            if lo is None:
                continue
            hi = sp.last_kept(dim.min, dim.max)
            new_min, new_max = sp.map_col(lo), sp.map_col(hi)
            letter = get_column_letter(new_min)
            # built by hand rather than copied: copy() drops the style array
            holder[letter] = ColumnDimension(
                ws, index=letter, min=new_min, max=new_max,
                width=dim.width, bestFit=dim.bestFit, hidden=dim.hidden,
                outlineLevel=dim.outlineLevel, collapsed=dim.collapsed,
                style=dim._style)

    def _apply_tables(self, ws, sp):
        sheet_key = _key(ws.title)
        for table in list(ws._tables.values()):
            plan = self.plan.tables[_key(table.name)]
            if plan.removed:
                del ws._tables[table.name]
                self._bump("tables removed")
                continue
            if sp.changed:
                # a table column's name is the header cell it sits over, so a
                # removed column takes its tableColumn with it
                cr = CellRange(table.ref)
                table.ref = self._shrink_range(table.ref, sp)
                table.tableColumns = [
                    column for offset, column in enumerate(table.tableColumns)
                    if not sp.is_removed(cr.min_col + offset)]
                if table.autoFilter is not None:
                    self._retarget_filter(table.autoFilter, sp)
                if table.sortState is not None:
                    self._retarget_sort(table, "sortState", sp)
                if plan.dropped_columns:
                    self._bump("tables narrowed")

            for column in table.tableColumns:
                for attr in ("calculatedColumnFormula", "totalsRowFormula"):
                    formula = getattr(column, attr)
                    if formula is None or not formula.attr_text:
                        continue
                    try:
                        formula.attr_text = self.rewriter.body(
                            formula.attr_text, sheet_key)
                    except _Poisoned:
                        setattr(column, attr, None)
                        self._bump("table formulas removed")

    def _retarget_filter(self, flt, sp):
        old = CellRange(str(flt.ref))
        new = self._shrink_range(old, sp)
        if new is None:
            flt.ref = None
            flt.filterColumn = []
            flt.sortState = None
            return False
        flt.ref = new
        new_min = CellRange(new).min_col
        columns = []
        for fc in flt.filterColumn:
            col = old.min_col + fc.colId
            if sp.is_removed(col):
                self._bump("filters removed")
                continue
            fc.colId = sp.map_col(col) - new_min
            columns.append(fc)
        flt.filterColumn = columns
        self._retarget_sort(flt, "sortState", sp)
        return True

    def _retarget_sort(self, owner, attr, sp):
        state = getattr(owner, attr)
        if state is None or state.ref is None:
            return
        new = self._shrink_range(str(state.ref), sp)
        if new is None:
            setattr(owner, attr, None)
            return
        state.ref = new
        conditions = []
        for cond in state.sortCondition:
            if cond.ref is None:
                continue
            ref = self._shrink_range(str(cond.ref), sp)
            if ref is None:
                continue
            cond.ref = ref
            conditions.append(cond)
        state.sortCondition = conditions

    def _apply_auto_filter(self, ws, sp):
        if not sp.changed or not ws.auto_filter:
            return
        if not self._retarget_filter(ws.auto_filter, sp):
            self._bump("filters removed")

    def _apply_validations(self, ws, sp):
        sheet_key = _key(ws.title)
        kept = []
        for dv in ws.data_validations.dataValidation:
            ranges = self._shrink_multi(dv.sqref.ranges, sp) if sp.changed \
                else list(dv.sqref.ranges)
            if not ranges:
                self._bump("validations removed")
                continue
            try:
                formulae = [self.rewriter.body(f, sheet_key)
                            for f in (dv.formula1, dv.formula2)]
            except _Poisoned:
                self._bump("validations removed")
                continue
            dv.formula1, dv.formula2 = formulae
            dv.sqref = MultiCellRange(ranges)
            kept.append(dv)
        ws.data_validations.dataValidation = kept

    def _apply_conditional_formats(self, ws, sp):
        sheet_key = _key(ws.title)
        old = ws.conditional_formatting
        new = ConditionalFormattingList()
        new.max_priority = old.max_priority
        for cf in old:
            ranges = self._shrink_multi(cf.sqref.ranges, sp) if sp.changed \
                else list(cf.sqref.ranges)
            if not ranges:
                self._bump("conditional formats removed")
                continue
            rules = []
            for rule in cf.rules:
                try:
                    formulae = [self.rewriter.body(f, sheet_key)
                                for f in (rule.formula or [])]
                    cfvo = self._rewrite_cfvo(rule, sheet_key)
                except _Poisoned:
                    self._bump("conditional formats removed")
                    continue
                rule.formula = formulae
                for obj, value in cfvo:
                    obj.val = value
                rules.append(rule)
            if not rules:
                continue
            holder = ConditionalFormatting(sqref=MultiCellRange(ranges),
                                           pivot=cf.pivot)
            new._cf_rules.setdefault(holder, []).extend(rules)
        ws.conditional_formatting = new

    def _rewrite_cfvo(self, rule, sheet_key):
        out = []
        for attr in ("colorScale", "dataBar", "iconSet"):
            part = getattr(rule, attr, None)
            if part is None:
                continue
            for obj in part.cfvo or ():
                if obj.type == "formula" and isinstance(obj.val, str):
                    out.append((obj, self.rewriter.body(obj.val, sheet_key)))
        return out

    def _apply_print_settings(self, ws, sp):
        if not sp.changed:
            return
        ranges = self._shrink_multi(ws._print_area.ranges, sp)
        area = PrintArea(ranges)
        area.title = ws._print_area.title
        ws._print_area = area

        cols = ws._print_cols
        if cols is not None:
            lo = column_index_from_string(cols.min_col.upper())
            hi = column_index_from_string(cols.max_col.upper())
            first = sp.first_kept(lo, hi)
            if first is None:
                ws._print_cols = None
            else:
                last = sp.last_kept(lo, hi)
                ws._print_cols = ColRange(
                    min_col=get_column_letter(sp.map_col(first)),
                    max_col=get_column_letter(sp.map_col(last)))

    def _apply_scenarios(self, ws, sp):
        if sp.changed and ws.scenarios:
            self._bump("scenarios removed", len(ws.scenarios.scenario))
            ws.scenarios = ScenarioList()
            return
        for scenario in ws.scenarios.scenario:
            if scenario.user:
                scenario.user = self.organisation

    def _apply_views(self, ws, sp):
        if not sp.changed:
            return
        for view in ws.views.sheetView:
            view.topLeftCell = self._safe_coord(view.topLeftCell, sp)
            pane = view.pane
            if pane is not None:
                if pane.xSplit:
                    split = int(pane.xSplit)
                    pane.xSplit = sum(1 for c in range(1, split + 1)
                                      if not sp.is_removed(c))
                pane.topLeftCell = self._safe_coord(pane.topLeftCell, sp)
                if not pane.xSplit and not pane.ySplit:
                    view.pane = None
                    view.selection = view.selection[:1]
            for sel in view.selection:
                sel.activeCell = self._safe_coord(sel.activeCell, sp) or "A1"
                sel.sqref = self._safe_sqref(sel.sqref, sp)

    def _safe_coord(self, coord, sp):
        if not coord:
            return coord
        try:
            return self._shrink_range(coord, sp)
        except (ValueError, TypeError):                   # pragma: no cover
            return "A1"

    def _safe_sqref(self, sqref, sp):
        if not sqref:
            return sqref
        out = []
        for part in str(sqref).split():
            try:
                new = self._shrink_range(part, sp)
            except (ValueError, TypeError):               # pragma: no cover
                new = None
            if new:
                out.append(new)
        return " ".join(out) or "A1"

    def _apply_breaks(self, ws, sp):
        if not sp.changed:
            return
        kept = []
        for brk in ws.col_breaks.brk:
            new = sp.map_col(brk.id) if brk.id else brk.id
            if new is None:
                continue
            brk.id = new
            kept.append(brk)
        ws.col_breaks.brk = kept

    # -- drawings ----------------------------------------------------------

    def _apply_charts(self, sheet):
        sheet_key = _key(sheet.title)
        sp = self.plan.sheets[sheet_key]
        kept = []
        for chart in getattr(sheet, "_charts", []):
            try:
                _walk_references(chart, self.rewriter, sheet_key)
            except _Poisoned:
                self._bump("charts removed")
                continue
            _strip_caches(chart)
            self._remap_anchor(chart, sp)
            kept.append(chart)
        if hasattr(sheet, "_charts"):
            sheet._charts = kept
        for image in getattr(sheet, "_images", []):
            self._remap_anchor(image, sp)

    def _nearest_kept(self, col, sp):
        """The new index of `col`, or of the first surviving column after it."""
        new = sp.map_col(col)
        if new is not None:
            return new
        following = sp.first_kept(col, max(sp.removed[-1] + 1, col))
        return sp.map_col(following) if following is not None else 1

    def _remap_anchor(self, owner, sp):
        anchor = getattr(owner, "anchor", None)
        if anchor is None or not sp.changed:
            return
        if isinstance(anchor, str):
            try:
                cr = CellRange(anchor)
            except (ValueError, TypeError):                # pragma: no cover
                return
            owner.anchor = "{0}{1}".format(
                get_column_letter(self._nearest_kept(cr.min_col, sp)),
                cr.min_row)
            return
        for attr in ("_from", "to"):
            marker = getattr(anchor, attr, None)
            if marker is None or not hasattr(marker, "col"):
                continue
            # anchor markers are zero-based
            marker.col = max(self._nearest_kept(marker.col + 1, sp) - 1, 0)

    # -- pivots ------------------------------------------------------------

    def _apply_pivots(self):
        for sheet in self.wb._sheets:
            pivots = getattr(sheet, "_pivots", None)
            if not pivots:
                continue
            if _key(sheet.title) in self.plan.dropped_sheets:
                continue                    # they go with the sheet
            host = self.plan.sheets[_key(sheet.title)]
            kept = []
            for pivot in pivots:
                if not host.changed and self._pivot_survives(pivot):
                    kept.append(pivot)
                else:
                    self._clear_pivot_output(sheet, pivot)
                    self._bump("pivot tables removed")
            sheet._pivots = kept

    @staticmethod
    def _clear_pivot_output(sheet, pivot):
        """
        The cells a pivot table renders into are a machine-made copy of its
        source rows, so they go with the pivot itself.
        """
        location = getattr(pivot, "location", None)
        ref = getattr(location, "ref", None)
        if not ref or not isinstance(sheet, Worksheet):
            return
        try:
            area = CellRange(ref)
        except (ValueError, TypeError):                   # pragma: no cover
            return
        for coord in area.cells:
            sheet._cells.pop(coord, None)

    def _pivot_survives(self, pivot):
        """
        A pivot cache is a verbatim copy of its source rows, and openpyxl
        cannot rebuild one.  Keep a pivot only when its source demonstrably
        never held any of the removed data.
        """
        cache = pivot.cache
        if cache is None:
            return False
        source = cache.cacheSource
        if source is None or source.type != "worksheet":
            return False
        wss = source.worksheetSource
        if wss is None:
            return False

        title = wss.sheet
        if title is None and wss.name:
            title = self._sheet_of_name(wss.name)
        if title is None and wss.ref:
            # a bare ref with no sheet is ambiguous
            return False
        sp = self.plan.sheets.get(_key(title)) if title else None
        if sp is None or sp.changed:
            return False
        for field in cache.cacheFields:
            if _key(field.name) in self.dropped_headers:
                return False
        return True

    def _sheet_of_name(self, name):
        defn = self.wb.defined_names.get(name)
        if defn is None:
            for sheet in self.wb._sheets:
                defn = sheet.defined_names.get(name)
                if defn is not None:
                    break
        if defn is None:
            return None
        try:
            for title, _cells in defn.destinations:
                return title
        except (ValueError, AttributeError, TokenizerError):
            return None
        return None

    # -- workbook ----------------------------------------------------------

    def _remove_sheets(self):
        for sheet in list(self.wb._sheets):
            if _key(sheet.title) in self.plan.dropped_sheets:
                self.wb.remove(sheet)
        if self.wb._sheets:
            self.wb._active_sheet_index = min(self.wb._active_sheet_index,
                                              len(self.wb._sheets) - 1)

    def _apply_properties(self):
        now = datetime.datetime.now(tz=datetime.timezone.utc).replace(tzinfo=None)
        self.wb.properties = DocumentProperties(creator=self.organisation,
                                                lastModifiedBy=self.organisation,
                                                created=now, modified=now)
        if len(self.wb.custom_doc_props):
            self._bump("custom properties removed",
                       len(self.wb.custom_doc_props))
            self.wb.custom_doc_props = CustomPropertyList()
        for sheet in self.wb._sheets:
            for pivot in getattr(sheet, "_pivots", []):
                if pivot.cache is not None and pivot.cache.refreshedBy:
                    pivot.cache.refreshedBy = self.organisation
        if self.plan.changed:
            self.wb.calculation.fullCalcOnLoad = True

    # -- reporting ---------------------------------------------------------

    @property
    def status(self):
        return STATUS_REDACTED if self.plan.changed else STATUS_UNCHANGED

    def detail(self):
        parts = []
        if self.removed_sheet_names:
            parts.append("sheets removed: "
                         + ", ".join(sorted(self.removed_sheet_names)))
        if self.removed_column_names:
            named = ["{0} ({1})".format(name, ", ".join(sheets))
                     for name, sheets in sorted(self.removed_column_names.items())]
            parts.append("columns removed: " + ", ".join(named))
        for what, count in sorted(self.counts.items()):
            parts.append(f"{what}: {count}")
        if not parts:
            return "nothing matched"
        return "; ".join(parts)

    def row(self, name):
        return {
            "workbook": name,
            "status": self.status,
            "sheets_removed": self.sheets_removed,
            "columns_removed": self.columns_removed,
            "cells_removed": self.cells_removed,
            "detail": self.detail(),
        }


def _runs(columns):
    """Group sorted column indices into (start, length) runs, right to left."""
    runs = []
    for col in columns:
        if runs and runs[-1][0] + runs[-1][1] == col:
            runs[-1][1] += 1
        else:
            runs.append([col, 1])
    return [(start, length) for start, length in reversed(runs)]


_CACHE_ATTRS = ("numCache", "strCache", "multiLvlStrCache")


def _children(obj, seen):
    """
    The attributes of a drawing object worth descending into, or None when
    `obj` is a leaf or leads back out of the drawing.
    """
    if obj is None or isinstance(obj, (str, bytes, int, float, bool)):
        return None
    # never follow a back-reference out into the sheet the chart sits on
    from openpyxl.workbook.child import _WorkbookChild
    from openpyxl.cell.cell import Cell
    from openpyxl.workbook.workbook import Workbook
    if isinstance(obj, (_WorkbookChild, Cell, Workbook)):
        return None
    attrs = getattr(obj, "__dict__", None)
    if not attrs or id(obj) in seen:
        return None
    seen.add(id(obj))
    return attrs


def _walk_references(obj, rewriter, sheet_key, seen=None):
    """Rewrite every `f` reference reachable from a chart object."""
    if seen is None:
        seen = set()
    if isinstance(obj, (list, tuple, set)):
        for item in obj:
            _walk_references(item, rewriter, sheet_key, seen)
        return
    attrs = _children(obj, seen)
    if attrs is None:
        return
    if isinstance(attrs.get("f"), str) and attrs["f"]:
        obj.f = rewriter.body(attrs["f"], sheet_key)
    for name, value in list(attrs.items()):
        if name != "f":
            _walk_references(value, rewriter, sheet_key, seen)


def _strip_caches(obj, seen=None):
    """Drop the cached copies of source values a chart carries around."""
    if seen is None:
        seen = set()
    if isinstance(obj, (list, tuple, set)):
        for item in obj:
            _strip_caches(item, seen)
        return
    attrs = _children(obj, seen)
    if attrs is None:
        return
    for name in _CACHE_ATTRS:
        if attrs.get(name) is not None:
            setattr(obj, name, None)
    for value in list(attrs.values()):
        _strip_caches(value, seen)


###############################################################################
# Public API
###############################################################################

def redact_workbook(source, destination, spec):
    """
    Write a redacted copy of the workbook at `source` to `destination`.

    :param source: path of the workbook to copy
    :param destination: path the copy is written to; it must not already exist
    :param spec: a mapping shaped like ``SPEC.json``
    :return: the report row for this workbook, as a dict keyed by
             :data:`REPORT_FIELDS`
    :raises SpecError: if the spec is not shaped correctly
    :raises RedactionError: if the copy could not be made; nothing is written
    """
    spec = validate_spec(spec)
    source = os.fspath(source)
    destination = os.fspath(destination)

    if os.path.exists(destination):
        raise RedactionError(
            f"a file is already at the destination {destination}")

    try:
        wb = load_workbook(source, read_only=False, keep_vba=False,
                           data_only=False, keep_links=True, rich_text=False)
    except RedactionError:                                # pragma: no cover
        raise
    except Exception as e:
        raise RedactionError(f"cannot read workbook: {e}") from e

    try:
        redactor = _WorkbookRedactor(wb, spec)
        redactor.plan_workbook()
        redactor.apply()
    except RedactionError:
        raise
    except Exception as e:
        raise RedactionError(f"cannot redact workbook: {e}") from e

    _save_new_file(wb, destination)
    return redactor.row(os.path.basename(source))


def _save_new_file(wb, destination):
    """
    Save `wb` at `destination`, leaving nothing behind if that cannot be
    finished and never replacing a file that is already there.
    """
    folder = os.path.dirname(destination) or "."
    try:
        handle, temporary = tempfile.mkstemp(dir=folder, prefix=".redact-",
                                             suffix=".part")
    except OSError as e:
        raise RedactionError(f"cannot write to {folder}: {e}") from e
    os.close(handle)
    try:
        try:
            wb.save(temporary)
        except Exception as e:
            raise RedactionError(f"cannot write copy: {e}") from e
        try:
            os.link(temporary, destination)
        except FileExistsError as e:
            raise RedactionError(
                f"a file is already at the destination {destination}") from e
        except (OSError, AttributeError, NotImplementedError):
            # hard links are not always available; fall back to a rename
            if os.path.exists(destination):
                raise RedactionError(
                    f"a file is already at the destination {destination}")
            try:
                os.replace(temporary, destination)
            except OSError as e:                          # pragma: no cover
                raise RedactionError(f"cannot write copy: {e}") from e
            return
    finally:
        if os.path.exists(temporary):
            try:
                os.remove(temporary)
            except OSError:                               # pragma: no cover
                pass


def _failed_row(name, message):
    return {
        "workbook": name,
        "status": STATUS_FAILED,
        "sheets_removed": 0,
        "columns_removed": 0,
        "cells_removed": 0,
        "detail": message,
    }


def find_workbooks(input_dir):
    """Every .xlsx file directly under `input_dir`, in name order."""
    try:
        names = os.listdir(input_dir)
    except OSError as e:
        raise RedactionError(f"cannot read {input_dir}: {e}") from e
    found = [n for n in names
             if n.lower().endswith(WORKBOOK_SUFFIX)
             and not n.startswith("~$")
             and os.path.isfile(os.path.join(input_dir, n))]
    return sorted(found)


def _is_inside(child, parent):
    child = os.path.realpath(child)
    parent = os.path.realpath(parent)
    return child == parent or child.startswith(parent + os.sep)


def redact_directory(input_dir, output_dir, spec, report=None):
    """
    Redact every workbook directly under `input_dir` into `output_dir`.

    :return: the list of report rows, one per workbook found
    """
    spec = validate_spec(spec)
    input_dir = os.fspath(input_dir)
    output_dir = os.fspath(output_dir)

    if not os.path.isdir(input_dir):
        raise RedactionError(f"{input_dir} is not a directory")
    if _is_inside(output_dir, input_dir):
        raise RedactionError(
            f"the output directory {output_dir} may not be inside {input_dir}")

    names = find_workbooks(input_dir)
    os.makedirs(output_dir, exist_ok=True)

    rows = []
    try:
        for name in names:
            source = os.path.join(input_dir, name)
            destination = os.path.join(output_dir, name)
            try:
                rows.append(redact_workbook(source, destination, spec))
            except Exception as e:
                rows.append(_failed_row(name, f"{type(e).__name__}: {e}"))
    finally:
        if report is not None:
            write_report(report, rows)
    return rows


def write_report(path, rows):
    """Write the CSV report for a run."""
    folder = os.path.dirname(os.fspath(path))
    if folder:
        os.makedirs(folder, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as out:
        writer = csv.DictWriter(out, fieldnames=list(REPORT_FIELDS))
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def summarise(rows):
    counts = {STATUS_REDACTED: 0, STATUS_UNCHANGED: 0, STATUS_FAILED: 0}
    for row in rows:
        counts[row["status"]] = counts.get(row["status"], 0) + 1
    return "redacted {0}, unchanged {1}, failed {2}".format(
        counts[STATUS_REDACTED], counts[STATUS_UNCHANGED], counts[STATUS_FAILED])


###############################################################################
# Command line
###############################################################################

def _parser():
    parser = argparse.ArgumentParser(
        prog="python -m openpyxl.redact",
        description="Write redacted copies of workbooks for an outside party.")
    parser.add_argument("input_dir", help="folder holding the workbooks")
    parser.add_argument("output_dir", help="folder the copies are written to")
    parser.add_argument("--spec", required=True,
                        help="JSON file naming the sheets and columns to drop")
    parser.add_argument("--report", default=None,
                        help="CSV file to write the run's report to")
    return parser


def main(argv=None):
    """Run the command line tool.  Returns the process exit status."""
    args = _parser().parse_args(argv)

    try:
        spec = load_spec(args.spec)
    except SpecError as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    try:
        rows = redact_directory(args.input_dir, args.output_dir, spec,
                                report=args.report)
    except RedactionError as e:
        print(f"error: {e}", file=sys.stderr)
        return 2
    except OSError as e:
        print(f"error: {e}", file=sys.stderr)
        return 2

    print(summarise(rows))
    return 1 if any(r["status"] == STATUS_FAILED for r in rows) else 0


if __name__ == "__main__":
    sys.exit(main())
