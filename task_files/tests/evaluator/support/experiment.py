"""Build a folder of workbooks, run the redaction command, observe what it did.

Every requirement observes outcomes only: the bytes and parts of the copies the
command wrote, what those copies say when opened with the library, the cells of
the report, the command's output and exit status, and whether the originals
changed. Nothing reads the submission's source.
"""

from __future__ import annotations

import csv
import datetime
import hashlib
import json
import os
import re
import shutil
import stat
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Iterable

from support.assertions import Evidence, not_evaluated

CANDIDATE_PYTHON = os.environ.get("CANDIDATE_PYTHON") or "python3"
CANDIDATE_REPO = os.environ.get("CANDIDATE_REPO") or "/workspace"
RUN_ROOT = "/tmp"
COMMAND_TIMEOUT = 300

# Values planted in the workbooks. Each carries a token that cannot appear in a
# redacted copy by accident.
SALARY_A = 50731
SALARY_B = 62842
SALARY_DUP = 71953
NI_A = "QQ123456C"
NI_B = "QQ654321D"
NI_DUP = "QQ777777E"
NOTES_TEXT = "NOTES-PHI-7731 salary review minutes"
AUTHOR = "Priya Kapoor"
MACHINE = "hr-laptop-17"
ORGANISATION = "Northwind Finance"
MARKER = "#REDACTED"
LINK_HOST = "hr-fileserver"
LINK_TARGET = "file:///\\\\hr-fileserver\\payroll\\bonus_2024.xlsx"
HIRE_DATE = datetime.date(2024, 3, 15)
EQ_TEXT = "=see HR before quoting"

DEFAULT_SPEC = {
    "drop_sheets": ["Notes"],
    "drop_columns": ["Salary", "NI Number", "owner"],
    "marker": MARKER,
    "organisation": ORGANISATION,
}

REPO_PIVOT_FIXTURE = Path(CANDIDATE_REPO) / "openpyxl" / "reader" / "tests" / "data" / "pivot.xlsx"
REPO_LINK_FIXTURE = Path(CANDIDATE_REPO) / "openpyxl" / "workbook" / "external_link" / "tests" / "data" / "book1.xlsx"


def _openpyxl():
    import openpyxl  # noqa: F401
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.table import Table
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill
    from openpyxl.worksheet.formula import ArrayFormula
    from openpyxl.utils.datetime import CALENDAR_MAC_1904
    from openpyxl.packaging.custom import StringProperty

    return dict(StringProperty=StringProperty, Workbook=Workbook, load_workbook=load_workbook, Comment=Comment,
                DefinedName=DefinedName, Table=Table, DataValidation=DataValidation,
                CellIsRule=CellIsRule, PatternFill=PatternFill, ArrayFormula=ArrayFormula,
                CALENDAR_MAC_1904=CALENDAR_MAC_1904)


class Batch:
    """A folder of workbooks with something to hide in every corner."""

    def __init__(
        self,
        *,
        staff: bool = True,
        dup: bool = False,
        dates1904: bool = False,
        pivot: bool = False,
        links: bool = False,
        plain: bool = True,
        locked: bool = False,
        non_workbook: bool = False,
        big: bool = False,
    ) -> None:
        o = _openpyxl()
        self.root = Path(tempfile.mkdtemp(prefix="redact", dir=RUN_ROOT))
        self.input_dir = self.root / "in"
        self.output_dir = self.root / "out"
        self.report = self.root / "report.csv"
        self.spec_path = self.root / "spec.json"
        self.input_dir.mkdir()
        self.spec_path.write_text(json.dumps(DEFAULT_SPEC), encoding="utf-8")
        self.files: dict[str, Path] = {}
        self.pivot_owner_values: list[str] = []

        def write(wb, name: str, key: str) -> None:
            path = self.input_dir / name
            wb.save(str(path))
            self.files[key] = path

        if staff:
            # One removed column here, so a run that gets the second deletion
            # wrong is measured on the payroll workbook and nowhere else.
            wb = o["Workbook"](); ws = wb.active; ws.title = "Staff"
            ws.append(["Name", "Salary", "Dept", "Grade", "Manager", "Bonus", "Note", "Ref"])
            ws.append(["Ann", SALARY_A, "Ops", "B", "Zed", "=B2*0.1", EQ_TEXT, "=E2"])
            ws["G2"].data_type = "s"  # a text that merely begins with '='
            ws.append(["Bob", SALARY_B, "Eng", "C", "Yara", "=B3*0.1", "ok", "=E3"])
            ws["A4"] = "Total"
            ws["B4"] = "=SUM(B2:B3)"          # depends on removed data
            ws["H4"] = "=COUNTA(E2:E3)"       # depends on kept data to the right of the removed column
            ws["I2"] = o["ArrayFormula"]("I2:I3", "=E2:E3")  # an array formula over kept data
            # formulas that name a column by text or by position rather than by reference
            ws["A8"] = '=INDIRECT("B2")'                 # the removed column, spelt as text
            ws["A9"] = '=INDIRECT("E2")'                 # a kept column, spelt as text
            ws["A10"] = "=OFFSET(A1,1,1)"                # the removed column, by position
            ws["A11"] = "=OFFSET(A1,1,4)"                # a kept column, by position
            ws["A12"] = '=VLOOKUP("Ann",A2:H3,2,FALSE)'  # column 2 of a range that loses column 2
            ws["A13"] = '=VLOOKUP("Ann",A2:H3,5,FALSE)'  # column 5 of a range that loses a column before it
            ws.merge_cells("A6:H6"); ws["A6"] = "Confidential - HR use"
            ws.add_table(o["Table"](displayName="StaffTable", ref="A1:H3"))
            dv = o["DataValidation"](type="whole", operator="greaterThan", formula1="0")
            dv.add("B2:B3"); ws.add_data_validation(dv)
            ws.conditional_formatting.add(
                "E2:E3", o["CellIsRule"](operator="equal", formula=['"Zed"'],
                                        fill=o["PatternFill"]("solid", fgColor="FFFF00")))
            ws.column_dimensions["E"].hidden = True
            ws.column_dimensions["C"].width = 22.0
            ws["E2"].comment = o["Comment"]("checked", AUTHOR)
            wb.defined_names["SalaryRange"] = o["DefinedName"]("SalaryRange", attr_text="Staff!$B$2:$B$3")
            wb.defined_names["Managers"] = o["DefinedName"]("Managers", attr_text="Staff!$E$2:$E$3")
            wb.defined_names["NoteTotal"] = o["DefinedName"]("NoteTotal", attr_text="Notes!$A$1")
            s = wb.create_sheet("Summary")
            s["A1"] = "=Staff!B2"              # removed column on another sheet
            s["A2"] = "=Staff!E2"              # kept column on another sheet, shifts
            s["A3"] = "=Notes!A1"              # removed sheet
            s["A4"] = "=SUM(Staff!B2:B3)"      # removed range on another sheet
            s["A5"] = "=Managers"              # a name that survives
            s["A6"] = "=SalaryRange"           # a name that pointed at removed data
            s["B1"] = 7
            n = wb.create_sheet("Notes"); n["A1"] = NOTES_TEXT
            wb.properties.creator = AUTHOR
            wb.properties.lastModifiedBy = MACHINE
            wb.properties.description = "Prepared by %s for the HR review" % AUTHOR
            wb.custom_doc_props.append(o["StringProperty"](name="Owner", value=AUTHOR))
            write(wb, "staff.xlsx", "staff")

        if dup:
            # The same header twice, one spelt loosely, and two removals far apart.
            wb = o["Workbook"](); ws = wb.active; ws.title = "Payroll"
            ws.append(["Name", "Salary", "Dept", "Grade", "NI Number", "Manager", " salary "])
            ws.append(["Cara", SALARY_A, "Ops", "A", NI_DUP, "Zed", SALARY_DUP])
            ws.append(["Dev", SALARY_B, "Eng", "B", NI_B, "Yara", SALARY_DUP + 1])
            ws.auto_filter.ref = "A1:G3"
            ws.auto_filter.add_filter_column(5, ["Zed"])   # a criterion on the Manager column
            write(wb, "dup.xlsx", "dup")

        if dates1904:
            wb = o["Workbook"](); wb.epoch = o["CALENDAR_MAC_1904"]
            ws = wb.active; ws.title = "Hires"
            ws.append(["Name", "Salary", "Start Date"])
            ws.append(["Eve", SALARY_A, HIRE_DATE])
            ws["C2"].number_format = "yyyy-mm-dd"
            write(wb, "dates1904.xlsx", "dates1904")

        if pivot and REPO_PIVOT_FIXTURE.is_file():
            wb = o["load_workbook"](str(REPO_PIVOT_FIXTURE))
            raw = wb["raw"]
            headers = [c.value for c in raw[1]]
            owner_col = headers.index("owner") + 1
            self.pivot_owner_values = sorted({str(raw.cell(row=r, column=owner_col).value)
                                              for r in range(2, raw.max_row + 1)
                                              if raw.cell(row=r, column=owner_col).value})
            for ws in wb.worksheets:
                for pv in ws._pivots:
                    src = pv.cache.cacheSource.worksheetSource
                    src.sheet = "raw"
                    src.ref = raw.dimensions
                    src.name = None
            write(wb, "campaigns.xlsx", "pivot")

        if links and REPO_LINK_FIXTURE.is_file():
            # The repository's external-link fixture, pointed at an internal server.
            target = self.input_dir / "budget.xlsx"
            with zipfile.ZipFile(str(REPO_LINK_FIXTURE)) as src, zipfile.ZipFile(str(target), "w", zipfile.ZIP_DEFLATED) as dst:
                for item in src.infolist():
                    data = src.read(item.filename)
                    if item.filename.endswith("externalLink1.xml.rels"):
                        data = data.replace(b'Target="book2.xlsx"', ('Target="%s"' % LINK_TARGET).encode("utf-8"))
                    dst.writestr(item, data)
            self.files["links"] = target

        if plain:
            wb = o["Workbook"](); ws = wb.active; ws.title = "Data"
            ws.append(["Item", "Count"]); ws.append(["widgets", 3])
            write(wb, "plain.xlsx", "plain")

        if locked:
            # An encrypted workbook is an OLE compound file, not a zip.
            path = self.input_dir / "00_locked.xlsx"
            path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 1024)
            self.files["locked"] = path

        if non_workbook:
            path = self.input_dir / "readme.txt"
            path.write_text("not a workbook; salary figures are in staff.xlsx\n", encoding="utf-8")
            self.files["non_workbook"] = path

        if big:
            wb = o["Workbook"](); ws = wb.active; ws.title = "Ledger"
            ws.append(["Name", "Salary", "Dept"] + ["c%d" % i for i in range(40)])
            for r in range(20000):
                ws.append(["p%d" % r, 1000 + r, "D%d" % (r % 7)] + list(range(40)))
            write(wb, "zz_big.xlsx", "big")

        self.input_digests = self.digests()

    # -- observing the inputs -------------------------------------------

    def digests(self) -> dict[str, str]:
        found = {}
        for path in sorted(self.input_dir.rglob("*")):
            if path.is_file():
                found[str(path.relative_to(self.input_dir))] = hashlib.sha256(path.read_bytes()).hexdigest()
        return found

    def workbook_names(self) -> list[str]:
        return sorted(p.name for p in self.input_dir.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx")

    def write_spec(self, spec: Any) -> Path:
        self.spec_path.write_text(json.dumps(spec), encoding="utf-8")
        return self.spec_path

    def close(self) -> None:
        shutil.rmtree(self.root, ignore_errors=True)


# ---------------------------------------------------------------------------
# Running the command
# ---------------------------------------------------------------------------


class Result:
    def __init__(self, completed: subprocess.CompletedProcess, batch: Batch,
                 output_dir: Path, report: Path) -> None:
        self.returncode = completed.returncode
        self.stdout = completed.stdout.decode("utf-8", "replace")
        self.stderr = completed.stderr.decode("utf-8", "replace")
        self.batch = batch
        self.output_dir = output_dir
        self.report = report

    def copies(self) -> list[Path]:
        if not self.output_dir.is_dir():
            return []
        return sorted(p for p in self.output_dir.rglob("*") if p.is_file())

    def copy_path(self, name: str) -> Path | None:
        for path in self.copies():
            if path.name == name:
                return path
        return None

    def workbook(self, name: str):
        path = self.copy_path(name)
        if path is None:
            return None
        import warnings
        o = _openpyxl()
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                return o["load_workbook"](str(path))
        except Exception:
            return None

    DATA_PARTS = ("xl/worksheets/", "xl/pivotCache/", "xl/pivotTables/", "xl/sharedStrings",
                  "xl/comments", "xl/tables/", "xl/externalLinks/", "xl/workbook.xml",
                  "docProps/", "xl/charts/", "xl/drawings/", "xl/threadedComments")

    def raw(self, name: str, prefixes: Iterable[str] | None = None) -> bytes:
        """The data-bearing parts of the copy, concatenated, for token searches.

        Theme and style parts are left out: they carry numbers of their own
        (tints, saturations) that have nothing to do with the cells.
        """
        path = self.copy_path(name)
        if path is None:
            return b""
        wanted = tuple(prefixes) if prefixes is not None else self.DATA_PARTS
        try:
            with zipfile.ZipFile(str(path)) as archive:
                return b"".join(archive.read(n) for n in archive.namelist()
                                if n.startswith(wanted))
        except zipfile.BadZipFile:
            return path.read_bytes()

    def parts(self, name: str) -> list[str]:
        path = self.copy_path(name)
        if path is None:
            return []
        try:
            with zipfile.ZipFile(str(path)) as archive:
                return archive.namelist()
        except zipfile.BadZipFile:
            return []

    def leaked(self, name: str, tokens: Iterable[Any], prefixes: Iterable[str] | None = None) -> list[str]:
        raw = self.raw(name, prefixes)
        return [str(t) for t in tokens if str(t).encode("utf-8") in raw]

    def report_rows(self) -> list[dict[str, str]]:
        if not self.report.is_file():
            return []
        with self.report.open(newline="", encoding="utf-8", errors="replace") as handle:
            rows = list(csv.reader(handle))
        if not rows:
            return []
        header = [h.strip() for h in rows[0]]
        return [dict(zip(header, row)) for row in rows[1:]]

    def report_header(self) -> list[str]:
        if not self.report.is_file():
            return []
        with self.report.open(newline="", encoding="utf-8", errors="replace") as handle:
            rows = list(csv.reader(handle))
        return [h.strip() for h in rows[0]] if rows else []

    def summary(self) -> dict[str, int] | None:
        for line in reversed(self.stdout.splitlines()):
            match = re.search(r"redacted\s+(\d+),\s*unchanged\s+(\d+),\s*failed\s+(\d+)", line)
            if match:
                return dict(zip(("redacted", "unchanged", "failed"), map(int, match.groups())))
        return None


def run_redact(
    batch: Batch,
    *,
    output_dir: Path | None = None,
    report: Path | None = None,
    spec_path: Path | None = None,
    extra: Iterable[str] = (),
    timeout: float = COMMAND_TIMEOUT,
) -> Result:
    output_dir = output_dir or batch.output_dir
    report = report or batch.report
    command = [CANDIDATE_PYTHON, "-m", "openpyxl.redact", str(batch.input_dir), str(output_dir),
               "--spec", str(spec_path or batch.spec_path), "--report", str(report), *extra]
    env = dict(os.environ)
    env.setdefault("PYTHONWARNINGS", "ignore")
    try:
        completed = subprocess.run(
            command, cwd=CANDIDATE_REPO, stdin=subprocess.DEVNULL, stdout=subprocess.PIPE,
            stderr=subprocess.PIPE, timeout=timeout, env=env,
            preexec_fn=lambda: os.umask(0o022),
        )
    except subprocess.TimeoutExpired as expired:
        completed = subprocess.CompletedProcess(
            command, 124, expired.stdout or b"", (expired.stderr or b"") + b"\n[timed out]")
    return Result(completed, batch, output_dir, report)


def run_redact_and_kill(batch: Batch, watch: str, *, timeout: float = COMMAND_TIMEOUT) -> Result:
    """Run the command and kill it the moment ``watch`` starts to exist in OUTPUT_DIR.

    A copy written straight to its destination is caught half-written; a copy
    written elsewhere and moved into place appears only when it is complete.
    """
    import signal
    import time

    command = [CANDIDATE_PYTHON, "-m", "openpyxl.redact", str(batch.input_dir), str(batch.output_dir),
               "--spec", str(batch.spec_path), "--report", str(batch.report)]
    env = dict(os.environ)
    env.setdefault("PYTHONWARNINGS", "ignore")
    target = batch.output_dir / watch
    process = subprocess.Popen(command, cwd=CANDIDATE_REPO, stdin=subprocess.DEVNULL,
                               stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env,
                               preexec_fn=lambda: os.umask(0o022))
    killed = False
    deadline = time.monotonic() + timeout
    while process.poll() is None and time.monotonic() < deadline:
        try:
            if target.exists() and target.stat().st_size > 0:
                process.kill()
                killed = True
                break
        except OSError:
            pass
        time.sleep(0.005)
    try:
        out, err = process.communicate(timeout=30)
    except subprocess.TimeoutExpired:
        process.kill()
        out, err = process.communicate()
    completed = subprocess.CompletedProcess(command, process.returncode, out or b"", err or b"")
    result = Result(completed, batch, batch.output_dir, batch.report)
    result.killed = killed  # type: ignore[attr-defined]
    return result


def run_python(code: str, timeout: float = 120) -> subprocess.CompletedProcess:
    return subprocess.run([CANDIDATE_PYTHON, "-c", code], cwd=CANDIDATE_REPO, stdin=subprocess.DEVNULL,
                          stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout)


# ---------------------------------------------------------------------------
# Guards shared by the safety requirements
# ---------------------------------------------------------------------------


_FEATURE: dict[str, Any] | None = None


def feature_probe() -> dict[str, Any]:
    batch = Batch(staff=True, plain=False)
    try:
        result = run_redact(batch)
        if result.copy_path("staff.xlsx") is None:
            return {"present": False,
                    "reason": "the redact command wrote no copy (exit %d): %s"
                              % (result.returncode, (result.stderr or result.stdout)[-400:])}
        return {"present": True, "reason": ""}
    finally:
        batch.close()


def require_feature(evidence: Evidence) -> None:
    global _FEATURE
    if _FEATURE is None:
        _FEATURE = feature_probe()
    evidence.add("feature_probe", _FEATURE.get("reason") or "present")
    if not _FEATURE.get("present"):
        not_evaluated("the redact command is not present: %s" % _FEATURE.get("reason"))


def require_copy(result: Result, name: str, evidence: Evidence):
    """The copy must exist and open before anything in it is judged."""
    evidence.add("exit", result.returncode)
    evidence.add("stdout_tail", result.stdout[-300:])
    evidence.add("stderr_tail", result.stderr[-600:])
    evidence.add("copies", [p.name for p in result.copies()])
    wb = result.workbook(name)
    if wb is None:
        not_evaluated("no readable copy of %s was written, so nothing in it can be judged" % name)
    return wb


def mode_of(path: Path) -> int | None:
    try:
        return stat.S_IMODE(path.stat().st_mode)
    except OSError:
        return None


def formula_text(cell) -> str | None:
    value = cell.value
    if value is None:
        return None
    text = getattr(value, "text", None)
    if text is not None:
        return str(text)
    return str(value)


def refers_to(text: str | None, *needles: str) -> bool:
    if not text:
        return False
    return any(n.casefold() in text.casefold() for n in needles)
